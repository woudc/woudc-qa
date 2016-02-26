# =================================================================
#
# Terms and Conditions of Use
#
# Unless otherwise noted, computer program source code of this
# distribution is covered under Crown Copyright, Government of
# Canada, and is distributed under the MIT License.
#
# The Canada wordmark and related graphics associated with this
# distribution are protected under trademark law and copyright law.
# No permission is granted to use them outside the parameters of
# the Government of Canada's corporate identity program. For
# more information, see
# http://www.tbs-sct.gc.ca/fip-pcim/index-eng.asp
#
# Copyright title to all 3rd party software distributed with this
# software is held by the respective copyright holders as noted in
# those files. Users are asked to read the 3rd Party Licenses
# referenced with those assets.
#
# Copyright (c) 2016 Government of Canada
#
# Permission is hereby granted, free of charge, to any person
# obtaining a copy of this software and associated documentation
# files (the "Software"), to deal in the Software without
# restriction, including without limitation the rights to use,
# copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following
# conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES
# OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT
# HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
# WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR
# OTHER DEALINGS IN THE SOFTWARE.
#
# =================================================================


import os
import logging
from collections import OrderedDict
import openpyxl
from woudc_extcsv import loads
from woudc_qa.util import \
    get_extcsv_value
from dataset_handlers import \
    OzoneSondeHandler, \
    TotalOzoneHandler

__version__ = '0.1.1'

__dirpath = os.path.dirname(os.path.realpath(__file__))

WOUDC_QA_RULES = os.path.join(__dirpath, 'woudc-qa-rules.xlsx')

LOGGER = logging.getLogger(__name__)


class QualityChecker(object):
    """Quality assess WOUDC data."""

    def __init__(self, extcsv, file_path, rule_def_path=None):
        """
        Quality assess incoming WOUDC data and maintain results.

        :param extcsv: woudc_extcsv Reader object
            containing WOUDC data to be qa'd
        """
        # create function to read qa-rules-definition and load all this
        self._file_path = file_path
        self._rule_path = None
        self._qa_rules = OrderedDict()
        self._extcsv = extcsv
        self._dataset = None
        try:
            self.dataset = \
                get_extcsv_value(self.extcsv, 'CONTENT', 'Category')
            self.dataset = self.dataset.lower()
        except Exception, err:
            msg = 'Unable to get CONTENT.Category. Due to: %s' % str(err)
            LOGGER.error(msg)
            raise err
        self._qa_flags = OrderedDict()
        self._qa_functions = OrderedDict()

        self._qa_results = OrderedDict()
        if self.file_path is None:
            self.file_path = 'file1'
        self.qa_results[self.file_path] = {}

        if rule_def_path is not None:
            self._rule_path = rule_def_path
        else:
            self._rule_path = WOUDC_QA_RULES

        try:
            self.execute()
        except Exception, err:
            msg = 'Unable to execute qa. Due to: %s' % str(err)
            LOGGER.critical(msg)
            raise err

    @property
    def qa_rules(self):
        """
        :returns: extcsv qa rule definitions
        """

        return self._qa_rules

    @qa_rules.setter
    def qa_rules(self, qa_rules):
        """
        Set qa rules
        """

        self._qa_rules = qa_rules

    @property
    def qa_flags(self):
        """
        :returns: extcsv qa flag definitions
        """

        return self._qa_flags

    @qa_flags.setter
    def qa_falgs(self, qa_flags):
        """
        Set qa flags
        """

        self._qa_flags = qa_flags

    @property
    def qa_functions(self):
        """
        :returns: extcsv qa function definitions
        """

        return self._qa_functions

    @qa_functions.setter
    def qa_functions(self, qa_functions):
        """
        Set qa functions
        """

        self._qa_functions = qa_functions

    @property
    def qa_results(self):
        """
        :returns: extcsv qa results
        """

        return self._qa_results

    @qa_results.setter
    def qa_results(self, qa_results):
        """
        Set qa results
        """

        self._qa_results = qa_results

    @property
    def extcsv(self):
        """
        :returns: extcsv
        """

        return self._extcsv

    @extcsv.setter
    def extcsv(self, extcsv):
        """
        Set extcsv
        """

        self._extcsv = extcsv

    @property
    def dataset(self):
        """
        :returns: dataset
        """

        return self._dataset

    @dataset.setter
    def dataset(self, dataset):
        """
        Set dataset
        """

        self._dataset = dataset

    @property
    def file_path(self):
        """
        :returns: file_path
        """

        return self._file_path

    @file_path.setter
    def file_path(self, file_path):
        """
        Set file_path
        """

        self._file_path = file_path

    @property
    def rule_path(self):
        """
        :returns: rule_path
        """

        return self._rule_path

    @rule_path.setter
    def rule_path(self, rule_path):
        """
        Set rule_path
        """

        self._rule_path = rule_path

    def execute(self):
        """
        orchestrate qa rules execution

        1) load up the rule definitions
        2) check precond
        3) check related tests
        4) run qa tests
        5) store qa results
        """
        # 1) load up the rule definitions
        try:
            self.load_qa_definitions()
        except Exception, err:
            msg = 'Unable to load definitions. Due to: %s' % str(err)
            LOGGER.critical(msg)
            raise err

        # get qa rules for this dataset
        if self.dataset not in self.qa_rules.keys():
            msg = 'No Qa rules defined for dataset: %s' % self.dataset
            LOGGER.error(msg)
            raise KeyError(msg)
        else:
            qa_rules = self.qa_rules[self.dataset]
            for rule in qa_rules:
                # check rule status
                rule_status = rule['test_status']
                if rule_status == '1':
                    result = None
                    continue_testing = False
                    profile = False
                    if rule['profile'] == '1':
                        profile = True
                    if rule['table_index'] == 'None':
                        rule['table_index'] = 1
                    # setup flag map
                    poss_results = rule['test_results'].split('|')
                    flag_map = {
                        True: None,
                        False: None,
                        'Error': 'Error'
                    }
                    if len(poss_results) == 2:
                        flag_map = {
                            True: poss_results[1],
                            False: poss_results[0]
                        }
                    if len(poss_results) == 1:
                        flag_map = {
                            True: poss_results[0]
                        }
                    # 2) check pre-condidtions
                    try:
                        result = self.check_preconditions(rule)
                    except Exception, err:
                        msg = 'Unable to run test_id: %s.\
                            Due to: preconditions unable to run.'\
                            % rule['test_id']
                        LOGGER.error(msg)
                        # if test fails to run, store NR for the test
                        result = 'NR'
                    # store result
                    try:
                        self._set_test_result(rule['test_id'], rule,
                                              'precond_result', result)
                    except Exception, err:
                        msg = 'Unable to set precondition test result.\
                        Due to: %s' % str(err)
                        LOGGER.error(msg)
                        continue
                    if any([result is None, result is True]):
                        continue_testing = True

                    # 2) check related test
                    if continue_testing:
                        result = None
                        continue_testing = False
                        # check if this rule is for profile field or not
                        # profile r tests needs to be run one per each row
                        row = 1
                        if not profile:
                            try:
                                result = self.check_related_test(rule, row)
                            except Exception, err:
                                msg = 'Unable to run test_id: %s.\
                                    Due to: related test unable to run.' % \
                                    rule['test_id']
                                LOGGER.error(msg)
                                result = 'NR'
                            # store result
                            try:
                                self._set_test_result(rule['test_id'], rule,
                                                      'related_test_result',
                                                      result)
                            except Exception, err:
                                msg = 'Unable to set related test result.\
                                Due to: %s' % str(err)
                                LOGGER.error(msg)
                                continue
                            if any([result is None, result is True]):
                                continue_testing = True
                        else:
                            continue_testing = True

                    # precond tests checked successfully
                    # related tests checked successfully (non-profile)
                    # check qa tests
                    if continue_testing:
                        result = None
                        # figure some stuff out
                        test_cate = rule['test_category']
                        # handle test categories
                        if test_cate == 'presence':
                            self.do_presence_check(rule, profile, flag_map)
                        elif test_cate == 'range':
                            self.do_range_check(rule, profile, flag_map)
                        elif test_cate == 'step':
                            self.do_step_check(rule, profile, flag_map)

    def do_step_check(self, rule, profile, flag_map):
        """
        do step check
        """
        result = None
        # unpackge rule
        table = rule['table']
        table_index = rule['table_index']
        field = rule['element']
        function = rule['function']
        param_a = rule['function_parameter_a']
        value = \
            get_extcsv_value(self.extcsv, table, field, table_index,
                             payload=profile)
        if profile:
            # get related tests
            row = 0
            val_len = len(value)
            while row < val_len - 3:
                continue_testing = False
                try:
                    # got to check row-1 and row
                    this_row_result = self.check_related_test(rule, row)
                    next_row_result = self.check_related_test(rule, row + 1)
                except Exception, err:
                    msg = 'Unable to run test_id: %s.\
                        Due to: related test unable to run.' % rule['test_id']
                    LOGGER.error(msg)
                    result = 'NR'
                # store result
                if all([this_row_result is True, next_row_result is True]):
                    result = True
                try:
                    self._set_test_result(
                        rule['test_id'],
                        rule,
                        'related_test_result',
                        result,
                        row + 1
                    )
                except Exception, err:
                    msg = 'Unable to set related test result.\
                    Due to: %s' % str(err)
                    LOGGER.error(msg)
                    row += 1
                    continue
                if any([result is None, result is True]):
                    continue_testing = True
                if continue_testing:
                    try:
                        t_result = None
                        # determine type of step check
                        if function == 'TS_0':
                            t_result =\
                                self._function_ts_0(value[row], value[row + 1],
                                                    param_a)
                        elif function == 'TS_2':
                            t_result =\
                                self._function_ts_2(value[row], value[row + 1],
                                                    param_a)
                        else:
                            msg = 'Unrecognized step check function: %s.\
                            for test_id: %s' % (function, rule['test_id'])
                            LOGGER.error(msg)
                            t_result = 'Error'
                        t_result = flag_map[t_result]
                    except Exception, err:
                        msg = 'Unable to do step check for test_id: %s. \
                            Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        t_result = 'Error'

                    try:
                        self._set_test_result(rule['test_id'], rule, 'result',
                                              t_result, row + 1)
                        if row == val_len - 2:
                            self._set_test_result(rule['test_id'], rule,
                                                  'result', t_result, row + 2)
                    except Exception, err:
                        msg = 'Unable to set test result for test id: %s \
                        Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        pass

                row += 1

    def do_range_check(self, rule, profile, flag_map):
        """
        do range check.
        """

        result = None
        # unpackge rule
        table = rule['table']
        table_index = rule['table_index']
        field = rule['element']
        function = rule['function']
        param_a = rule['function_parameter_a']
        param_b = rule['function_parameter_b']
        # get value from extcsv
        value = \
            get_extcsv_value(self.extcsv, table, field, table_index,
                             payload=profile)
        if profile:
            # get related tests
            row = 1
            for val in value:
                continue_testing = False
                try:
                    result = self.check_related_test(rule, row)
                except Exception, err:
                    msg = 'Unable to run test_id: %s.\
                        Due to: related test unable to run.' % rule['test_id']
                    LOGGER.error(msg)
                    result = 'NR'
                # store result
                try:
                    self._set_test_result(rule['test_id'], rule,
                                          'related_test_result', result, row)
                except Exception, err:
                    msg = 'Unable to set related test result.\
                    Due to: %s' % str(err)
                    LOGGER.error(msg)
                    row += 1
                    continue
                if any([result is None, result is True]):
                    continue_testing = True
                if continue_testing:
                    try:
                        t_result = None
                        # determine type of range check
                        if function == 'RC_1':
                            t_result = \
                                self._function_rc_1(param_a, param_b, val)
                        elif function == 'RC_5':
                            t_result = self._function_rc_5(param_a, val)
                        elif function == 'RC_6':
                            t_result = self._function_rc_6(param_a, val)
                        else:
                            msg = 'Unrecognized range check function: %s.\
                                for test_id: %s' % (function, rule['test_id'])
                            LOGGER.error(msg)
                            t_result = 'Error'
                        t_result = flag_map[t_result]
                    except Exception, err:
                        msg = 'Unable to do range check for test_id: %s. \
                            Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        t_result = 'Error'
                    try:
                        self._set_test_result(rule['test_id'], rule, 'result',
                                              t_result, row)
                    except Exception, err:
                        msg = 'Unable to set test result for test id: %s \
                            Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        continue
                row += 1
        else:
            try:
                t_result = None
                # determine type of range check
                if function == 'RC_1':
                    t_result = self._function_rc_1(param_a, param_b, value)
                elif function == 'RC_5':
                    t_result = self._function_rc_5(param_a, value)
                elif function == 'RC_6':
                    t_result = self._function_rc_6(param_a, value)
                else:
                    msg = 'Unrecognized range check function: %s.\
                    for test_id: %s' % (function, rule['test_id'])
                    LOGGER.error(msg)
                    t_result = 'Error'
                t_result = flag_map[t_result]
            except Exception, err:
                msg = 'Unable to do range check for test_id: %s. Due to: %s' \
                      % (rule['test_id'], str(err))
                LOGGER.error(msg)
                t_result = 'Error'
                # continue
            try:
                self._set_test_result(rule['test_id'], rule, 'result',
                                      t_result, 1)
            except Exception, err:
                msg = 'Unable to set test result for test id: %s \
                Due to: %s' % (rule['test_id'], str(err))
                pass
                LOGGER.error(msg)

    def do_presence_check(self, rule, profile, flag_map):
        """
        do presence check.
        """

        result = None
        # unpackge rule
        table = rule['table']
        table_index = rule['table_index']
        field = rule['element']
        function = rule['function']
        # get value from extcsv
        value = \
            get_extcsv_value(self.extcsv, table, field, table_index,
                             payload=profile)
        if profile:
            # get related tests
            row = 1
            for val in value:
                continue_testing = False
                try:
                    result = self.check_related_test(rule, row)
                except Exception, err:
                    msg = 'Unable to run test_id: %s.\
                        Due to: related test unable to run.' % rule['test_id']
                    LOGGER.error(msg)
                    result = 'NR'
                # store result
                try:
                    self._set_test_result(rule['test_id'], rule,
                                          'related_test_result', result, row)
                except Exception, err:
                    msg = 'Unable to set related test result.\
                    Due to: %s' % str(err)
                    LOGGER.error(msg)
                    row += 1
                    continue
                if any([result is None, result is True]):
                    continue_testing = True
                if continue_testing:
                    try:
                        t_result = None
                        if function == 'PR_1':
                            t_result = self._function_pc_1(val)
                        else:
                            msg = 'Unrecognized presence check function: %s\
                            in test_id: %s' % (function, rule['test_id'])
                            LOGGER.error(msg)
                            t_result = 'Error'
                        t_result = flag_map[t_result]
                    except Exception, err:
                        msg = 'Unable to do presence check for test_id: %s. \
                            Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        t_result = 'Error'
                    try:
                        self._set_test_result(rule['test_id'], rule, 'result',
                                              t_result, row)
                    except Exception, err:
                        msg = 'Unable to set test result for test id: %s \
                        Due to: %s' % (rule['test_id'], str(err))
                        LOGGER.error(msg)
                        pass
                row += 1
        else:
            try:
                t_result = None
                if function == 'PR_1':
                    t_result = self._function_pc_1(value)
                else:
                    msg = 'Unrecognized presence check function: %s\
                    in test_id: %s' % (function, rule['test_id'])
                    LOGGER.error(msg)
                    t_result = 'Error'
                t_result = flag_map[t_result]
            except Exception, err:
                msg = 'Unable to do presence check for test_id: %s. \
                    Due to: %s' % (rule['test_id'], str(err))
                LOGGER.error(msg)
                t_result = 'Error'
            try:
                self._set_test_result(rule['test_id'], rule, 'result',
                                      t_result, 1)
            except Exception, err:
                msg = 'Unable to set test result for test id: %s \
                Due to: %s' % (rule['test_id'], str(err))
                LOGGER.error(msg)
                pass

    def load_qa_definitions(self):
        """
        Load qa rules, functions and flag definitions
        """

        # load xlsx
        try:
            qa_defs_xl = \
                openpyxl.load_workbook(self.rule_path, use_iterators=True)
        except Exception, err:
            msg = 'Unable to read qa definition xlsx. Due to: %s' % str(err)
            LOGGER.critical(msg)
            raise err

        # load qa rules
        try:
            qa_tests_ws = qa_defs_xl.get_sheet_by_name('qa tests')
        except Exception, err:
            msg = 'Unable to get workshet: qa tests. Due to: %s' % str(err)
            LOGGER.critical(msg)
            raise err
        max_rows = qa_tests_ws.max_row + 1
        max_cols = qa_tests_ws.max_column + 1

        header = []
        try:
            for i in range(1, max_rows):
                dataset = None
                rule = {}
                for j in range(1, max_cols):
                    val = str(qa_tests_ws.cell(row=i, column=j).value)
                    if i == 1:  # header
                        header.append(val)
                    else:
                        if j == 1:  # dataset
                            dataset = val
                            if dataset not in self.qa_rules.keys():
                                self.qa_rules[dataset] = []
                        else:
                            rule_tok = header[j - 1]
                            if rule_tok not in rule.keys():
                                rule[rule_tok] = val
                if len(rule) != 0:
                    self.qa_rules[dataset].append(rule)
        except Exception, err:
            msg = 'Error while parsing qa tests tab: %s' % str(err)
            LOGGER.error(msg)
            raise err

        # load function defs
        try:
            func_ws = qa_defs_xl.get_sheet_by_name('function defs')
        except Exception, err:
            msg = 'Unable to get workshet: function defs. Due to: %s' \
                % str(err)
            LOGGER.critical(msg)
            raise err
        max_rows = func_ws.max_row + 1
        max_cols = func_ws.max_column + 1
        try:
            for i in range(2, max_rows):
                if func_ws.cell(row=i, column=1):
                    func = str(func_ws.cell(row=i, column=1).value)
                    func_def = str(func_ws.cell(row=i, column=2).value)
                    self.qa_functions[func] = func_def
        except Exception, err:
            msg = 'Error while parsing function defs tab: %s' % str(err)
            LOGGER.error(msg)
            raise err

        # load flag defs
        try:
            flag_ws = qa_defs_xl.get_sheet_by_name('flag defs')
        except Exception, err:
            msg = 'Unable to get workshet: qa flag defs. Due to: %s' % str(err)
            LOGGER.critical(msg)
            raise err
        max_rows = flag_ws.max_row + 1
        max_cols = flag_ws.max_column + 1
        try:
            for i in range(2, max_rows):
                if flag_ws.cell(row=i, column=1):
                    flag_val = flag_ws.cell(row=i, column=1).value
                    flag_name = flag_ws.cell(row=i, column=2).value
                    flag_desc = flag_ws.cell(row=i, column=3).value
                    self.qa_flags[flag_val] = {
                        'flag_name': flag_name,
                        'flag_desc': flag_desc
                    }
        except Exception, err:
            msg = 'Error while parsing function defs tab: %s' % str(err)
            LOGGER.error(msg)
            raise err

    def check_related_test(self, rule, row):
        """
        check related test

        :param rule: rule tokens
        :returns: boolean (pass/fail) or None (unable to check)
        """

        result = None
        if all(['related_test_id' not in rule.keys(),
                'related_test_status' not in rule.keys()]):
            return result
        else:
            # retrieve related test result
            r_test_id = rule['related_test_id'].split(',')
            r_test_result = rule['related_test_result'].split(',')
            if any([r_test_id == ['None'], len(r_test_id) == 0,
                    r_test_result == ['None'], len(r_test_result) == 0]):
                return result
            try:
                result = None
                for rtid in r_test_id:
                    if self._get_rule(rtid.strip(), 'profile') == '0':
                        row = 1
                    result = self._get_test_result(rtid, row)
                    rtr = r_test_result[r_test_id.index(rtid)].strip()
                    if result == (rtr):
                        result = True
                    else:
                        return False
            except Exception, err:
                LOGGER.error(str(err))
                raise err

        return result

    def check_preconditions(self, rule):
        """
        check preconditions
        if all precond are met, return True else False

        :param rule: rule package
        :returns: boolean or None (unable to check)
        """
        # store precondition results
        result = {
            'agency': None,
            'platform': None,
            'instrument_type': None,
            'instrument_model': None,
            'instrument_serial': None,
            'instrument_lat': None,
            'instrument_lon': None
        }
        # agency
        if 'agency' not in rule.keys():
            result.pop('agency')
        else:
            agency = rule['agency']
            if any([agency == 'None', agency == '']):
                result.pop('agency')
            else:
                try:
                    agency_f = \
                        get_extcsv_value(self.extcsv, 'DATA_GENERATION',
                                         'Agency')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['agency'] = None
                if agency == agency_f:
                    result['agency'] = True
                else:
                    result['agency'] = False
        # platform
        if 'platform' not in rule.keys():
            result.pop('platform')
        else:
            platform = rule['platform']
            if any([platform == '', platform == 'None']):
                result.pop('platform')
            else:
                try:
                    p_type_f = \
                        get_extcsv_value(self.extcsv, 'PLATFORM', 'Type')
                    p_id_f = get_extcsv_value(self.extcsv, 'PLATFORM', 'ID')
                    p_f = '%s%s' % (p_type_f, p_id_f)
                    p_f = p_f.lower()
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['platform'] = None
                if platform.lower() == p_f:
                    result['platform'] = True
                else:
                    result['platform'] = False
        # instrument type
        if 'instrument_type' not in rule.keys():
            result.pop('instrument_type')
        else:
            instrument_type = rule['instrument_type']
            if any([instrument_type == '', instrument_type == 'None']):
                result.pop('instrument_type')
            else:
                try:
                    i_type_f = \
                        get_extcsv_value(self.extcsv, 'INSTRUMENT', 'Name')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['instrument_type'] = None
                if instrument_type.lower() == i_type_f.lower():
                    result['instrument_type'] = True
                else:
                    result['instrument_type'] = False
        # instrument model
        if 'instrument_model' not in rule.keys():
            result.pop('instrument_model')
        else:
            instrument_model = rule['instrument_model']
            if any([instrument_model == '', instrument_model == 'None']):
                result.pop('instrument_model')
            else:
                try:
                    i_model_f = \
                        get_extcsv_value(self.extcsv, 'INSTRUMENT', 'Model')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['instrument_model'] = None
                if instrument_model.lower() == i_model_f.lower():
                    result['instrument_model'] = True
                else:
                    result['instrument_model'] = False
        # instrument serial number
        if 'instrument_serial_number' not in rule.keys():
            result.pop('instrument_serial')
        else:
            instrument_serial_number = rule['instrument_serial_number']
            if any([instrument_serial_number == '',
                    instrument_serial_number == 'None']):
                result.pop('instrument_serial')
            else:
                try:
                    i_num_f = \
                        get_extcsv_value(self.extcsv, 'INSTRUMENT', 'Number')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['instrument_serial'] = None
                if instrument_serial_number.lower() == i_num_f.lower():
                    result['instrument_serial'] = True
                else:
                    result['instrument_serial'] = False
        # instrument latitude
        if 'instrument_latitude' not in rule.keys():
            result['instrument_lat'] = None
            result.pop('instrument_lat')
        else:
            instrument_latitude = rule['instrument_latitude']
            if any([instrument_latitude == '', instrument_latitude == 'None']):
                result.pop('instrument_lat')
            else:
                try:
                    lat_f = \
                        get_extcsv_value(self.extcsv, 'LOCATION', 'Latitude')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['instrument_lat'] = None
                if ',' in instrument_latitude:
                    a, b = instrument_latitude.split(',')
                    r = self._function_rc_1(a, b, lat_f)
                    if r:
                        result['instrument_lat'] = True
                    else:
                        result['instrument_lat'] = False
                else:
                    if instrument_latitude == lat_f:
                        result['instrument_lat'] = True
                    else:
                        result['instrument_lat'] = False
        # instrument longitude
        if 'instrument_longitude' not in rule.keys():
            result.pop('instrument_lon')
        else:
            instrument_longitude = rule['instrument_longitude']
            if any([instrument_longitude == '',
                    instrument_longitude == 'None']):
                result.pop('instrument_lon')
            else:
                try:
                    lon_f = \
                        get_extcsv_value(self.extcsv, 'LOCATION', 'Longitude')
                except Exception, err:
                    msg = str(err)
                    LOGGER.error(msg)
                    result['instrument_lon'] = None
                if ',' in instrument_longitude:
                    a, b = instrument_longitude.split(',')
                    r = self._function_rc_1(a, b, lon_f)
                    if r:
                        result['instrument_lon'] = True
                    else:
                        result['instrument_lon'] = False
                else:
                    if instrument_longitude == lon_f:
                        result['instrument_lon'] = True
                    else:
                        result['instrument_lon'] = False

        v = result.values()
        if len(v) == 0:
            return None
        elif False in v:
            return False
        elif None in v:
            return None
        else:
            return True

    def _get_test_result(self, test_id, row):
        """
        helper method: retrieve test result

        :param test_id: test_id result to be retrieved
        :param row: row number of the element
        :returns: test result or None if test is n/a
        """
        result = None
        try:
            if test_id in self.qa_results[self.file_path].keys():
                if row in self.qa_results[self.file_path][test_id].keys():
                    result = \
                        self.qa_results[self.file_path][test_id][row]['result']
            else:
                return None
        except Exception, err:
            msg = 'Unable to get related test result for test_id: %s, \
                   row: %s. Due to: %s' % (test_id, row, str(err))
            LOGGER.error(msg)
            raise err

        return result

    def _set_test_result(self, test_id, rule, test_tok, result, row=1):
        """
        helper method: set qa test result

        :param test_id: test_id to set
        :param row: row number for which this test result applies
        """
        try:
            if test_id not in self.qa_results[self.file_path].keys():
                self.qa_results[self.file_path][test_id] = OrderedDict()
                if row not in self.qa_results[self.file_path][test_id].keys():
                    self.qa_results[self.file_path][test_id][row] = {
                        'result': None,
                        'table': rule['table'],
                        'table_index': rule['table_index'],
                        'element': rule['element'],
                        'related_test_id': rule['related_test_id'],
                        'related_test_result': None,
                        'precond_result': None,
                    }
                    self.qa_results[self.file_path][test_id]['test_def'] = rule
                    self.qa_results[self.file_path][test_id][row][test_tok] = \
                        result
                else:
                    self.qa_results[self.file_path][test_id][row][test_tok] = \
                        result
            else:
                if row not in self.qa_results[self.file_path][test_id].keys():
                    self.qa_results[self.file_path][test_id][row] = {
                        'result': None,
                        'table': rule['table'],
                        'table_index': rule['table_index'],
                        'element': rule['element'],
                        'related_test_id': rule['related_test_id'],
                        'related_test_result': None,
                        'precond_result': None,
                    }
                self.qa_results[self.file_path][test_id][row][test_tok] = \
                    result
        except Exception, err:
            msg = 'Unable to set test result. Due to: %s' % str(err)
            LOGGER.error(msg)
            raise err

    def test_definition_validation(self):
        """
        validate test definition provided in xlsx
        """
        return

    def _function_pc_1(self, value):
        """
        checks value is non empty
        """

        if all([value is not None, value != '']):
            return True
        else:
            return False

    def _function_rc_1(self, a, b, x):
        """
        evaluate a <= x <= b
        """

        try:
            a_f = float(a)
            b_f = float(b)
            x_f = float(x)
        except Exception, err:
            msg = str(err)
            LOGGER.error(msg)
            return 'Error'

        return a_f <= x_f <= b_f

    def _function_rc_5(self, a, x):
        """
        evaluate a <= x
        """

        try:
            a_f = float(a)
            x_f = float(x)
        except Exception, err:
            msg = str(err)
            LOGGER.error(msg)
            return 'Error'

        return a_f <= x_f

    def _function_rc_6(self, a, x):
        """
        evaluate a >= x
        """

        try:
            a_f = float(a)
            x_f = float(x)
        except Exception, err:
            msg = str(err)
            LOGGER.error(msg)
            return 'Error'

        return a_f >= x_f

    def _function_ts_0(self, a, b, x):
        """
        evaluable
        | a - b | = x
        """

        try:
            a_f = float(a)
            b_f = float(a)
            x_f = float(x)
        except Exception, err:
            msg = str(err)
            LOGGER.error(msg)
            return 'Error'

        return abs(a_f - b_f) == x_f

    def _function_ts_2(self, a, b, x):
        """
        evaluable
        | a - b | <= x
        """

        try:
            a_f = float(a)
            b_f = float(b)
            x_f = float(x)
        except Exception, err:
            msg = str(err)
            LOGGER.error(msg)
            return 'Error'

        return abs(a_f - b_f) <= x_f

    def _get_rule(self, test_id, rule_tok=None):
        """
        helper method:
        returns rule package at test_id
        """

        for rule in self.qa_rules[self.dataset]:
            if rule['test_id'] == test_id:
                if rule_tok is not None:
                    return rule[rule_tok]
                else:
                    return rule


class WOUDCQaExecutionError(Exception):
    """WOUDC quality assessment execution error"""

    def __init__(self, message, errors):
        """provide an error message and error stack"""
        super(WOUDCQaExecutionError, self).__init__(message)
        self.errors = errors


def qa(file_content, file_path=None, rule_path=None):
    """
    Parse incoming file content, invoke dataset handlers,
    and invoke quality checker

    :param file_content: file as string
    :param file_path: path to file (optional)
    """

    # parse incoming file content
    try:
        ecsv = loads(file_content)
    except Exception, err:
        msg = 'Unable to parse file. Due to: %s' % str(err)
        LOGGER.error(msg)
        raise err

    # figue our dataset
    dataset = get_extcsv_value(
        ecsv,
        'CONTENT',
        'Category'
    )

    # invoke dataset handler
    dataset_handler = None
    try:
        if dataset.lower() == 'ozonesonde':
            dataset_handler = OzoneSondeHandler(ecsv)
        if dataset.lower() == 'totalozone':
            dataset_handler = TotalOzoneHandler(ecsv)
    except Exception, err:
        msg = 'No handler found for dataset: %s. Cannot continue.' % \
            dataset.lower()
        LOGGER.critical(msg)
        raise err
    # invoke quality checker
    try:
        qa_checker = QualityChecker(
            dataset_handler.extcsv,
            file_path,
            rule_path
        )
    except Exception, err:
        msg = 'Unable to run Qa. Due to: %s' % str(err)
        LOGGER.critical(msg)
        raise err

    return qa_checker.qa_results
